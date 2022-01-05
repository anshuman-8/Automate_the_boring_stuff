import openpyxl

wb=openpyxl.load_workbook('multiplication.xlsx')
sheet=wb.active()

for r in range(1,11):
    sheet.cell(row=r, column=0).values=r
    sheet.cell(row=0, column=r).values=r

for r in range(1,11):
    for c in range(1,11):
        sheet.cell(row=r, column=c).values=r*c

